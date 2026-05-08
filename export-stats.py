"""
Export Stats.xlsx to JSON files for the F1 Clash Optimizer v3.5.
Usage: py tools/clash/export-stats.py

Generates:
  data/components.json  — All component stats (with overtakeMode for Battery)
  data/drivers.json     — All driver stats
  data/ai_compare.json  — AI Compare data for Series races only
  data/gp_compare.json  — AI Compare data for Gran Prix (Junior/Challenger/Contender/Champions)
  data/boosts.json      — Boost items and their stat bonuses
"""
import json
import os
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(SCRIPT_DIR, 'Stats.xlsx')
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

# --- COMPONENTS ---
# Columns: A=Code Name, B=Name, C=Level, D=Speed, E=Cornering, F=Power Unit, G=Qualifying,
#          H=Pit Time, I=Overtake Mode (total), J=Impact, K=Duration, L=Recharge,
#          M=Total, N=Team Score, O=Series, P=Rarity, Q-T=Cards/Coins, U=Component
ws = wb['Components']
rows = list(ws.iter_rows(min_row=1, values_only=True))

components = []
for row in rows[1:]:
    if not row[1]:
        continue
    name = str(row[1])
    level = int(row[2]) if row[2] else 0
    if level <= 0:
        continue
    comp_type = str(row[20] or '')
    components.append({
        "name": name,
        "level": level,
        "speed": float(row[3] or 0),
        "cornering": float(row[4] or 0),
        "powerUnit": float(row[5] or 0),
        "qualifying": float(row[6] or 0),
        "pitTime": float(row[7] or 0),
        "overtakeMode": float(row[8] or 0) if comp_type == 'Battery' else 0,
        "series": int(row[14] or 0),
        "rarity": str(row[15] or ''),
        "type": comp_type
    })

with open(os.path.join(DATA_DIR, 'components.json'), 'w') as f:
    json.dump(components, f, indent=1)
print(f"Components: {len(components)} entries written")

# --- DRIVERS ---
# Columns: A=link, B=Name, C=Rarity, D=Level, E=overtaking, F=defending, G=qualifying,
#          H=raceStart, I=tyreUse, J=Total, K=Team Score, L=Series
ws = wb['Drivers']
rows = list(ws.iter_rows(min_row=1, values_only=True))

drivers = []
for row in rows[1:]:
    if not row[1]:
        continue
    name = str(row[1])
    level = int(row[3]) if row[3] else 0
    if level <= 0:
        continue
    drivers.append({
        "name": name,
        "rarity": str(row[2] or ''),
        "level": level,
        "overtaking": float(row[4] or 0),
        "defending": float(row[5] or 0),
        "qualifying": float(row[6] or 0),
        "raceStart": float(row[7] or 0),
        "tyre": float(row[8] or 0),
        "total": float(row[9] or 0),
        "series": int(row[11] or 0)
    })

with open(os.path.join(DATA_DIR, 'drivers.json'), 'w') as f:
    json.dump(drivers, f, indent=1)
print(f"Drivers: {len(drivers)} entries written")

# --- AI COMPARE (new format: flat table with Location column) ---
# Columns: A=Tag, B=Location, C=Team, D=Driver, E=Overtaking, F=Defending, G=Qualifying,
#          H=Race Start, I=Tyre Use, J=Speed, K=Cornering, L=Power Unit, M=Qualifying(car),
#          N=Pit Time, O=Team Score
ws = wb['AI Compare']
rows = list(ws.iter_rows(min_row=1, values_only=True))

GP_KEYWORDS = ['Junior', 'Challenger', 'Contender', 'Champions']
GP_EXCLUDE = ['QA']  # Exclude test/QA entries

ai_compare = {}  # { "Series N": [...] } — series races only
gp_compare = {}  # { "Location": [...] } — GP events only

for row in rows[1:]:
    if not row[0] or not row[1]:
        continue
    location = str(row[1]).strip()
    driver = str(row[3] or '').strip()
    if not driver:
        continue

    entry = {
        "team": str(row[2] or ''),
        "driver": driver,
        "overtaking": float(row[4] or 0),
        "defending": float(row[5] or 0),
        "qualifying": float(row[6] or 0),
        "raceStart": float(row[7] or 0),
        "tyreMgmt": float(row[8] or 0),
        "speed": float(row[9] or 0),
        "cornering": float(row[10] or 0),
        "powerUnit": float(row[11] or 0),
        "qualifyingCar": float(row[12] or 0),
        "pitTime": float(row[13] or 0),
        "teamScore": float(row[14] or 0)
    }

    # Determine if this is a Series race or GP event
    is_gp = any(kw in location for kw in GP_KEYWORDS)
    is_excluded = any(ex in location for ex in GP_EXCLUDE)

    if is_excluded:
        continue  # Skip QA/test entries

    if is_gp:
        if location not in gp_compare:
            gp_compare[location] = []
        gp_compare[location].append(entry)
    else:
        # Series data — use location as key (e.g., "Series 4")
        if location not in ai_compare:
            ai_compare[location] = []
        ai_compare[location].append(entry)

with open(os.path.join(DATA_DIR, 'ai_compare.json'), 'w') as f:
    json.dump(ai_compare, f, indent=1)
ai_total = sum(len(v) for v in ai_compare.values())
print(f"AI Compare (Series): {len(ai_compare)} locations, {ai_total} entries written")

with open(os.path.join(DATA_DIR, 'gp_compare.json'), 'w') as f:
    json.dump(gp_compare, f, indent=1)
gp_total = sum(len(v) for v in gp_compare.values())
print(f"GP Compare: {len(gp_compare)} locations, {gp_total} entries written")

# --- BOOSTS ---
# Columns: A=Name, B=Overtaking, C=Defending, D=Race Start, E=Tyre Use, F=Speed,
#          G=Cornering, H=Power Unit, I=Pit Time, J=Impact, K=Duration, L=Recharge, M=Total
ws = wb['Boosts']
rows = list(ws.iter_rows(min_row=1, values_only=True))

boosts = []
for row in rows[1:]:
    if not row[0]:
        continue
    name = str(row[0]).strip()
    # Skip header rows
    if name == 'Name' or name == 'Overtaking':
        continue
    try:
        boosts.append({
        "name": name,
        "overtaking": float(row[1] or 0),
        "defending": float(row[2] or 0),
        "raceStart": float(row[3] or 0),
        "tyre": float(row[4] or 0),
        "speed": float(row[5] or 0),
        "cornering": float(row[6] or 0),
        "powerUnit": float(row[7] or 0),
        "pitTime": float(row[8] or 0),
        "impact": float(row[9] or 0),
        "duration": float(row[10] or 0),
        "recharge": float(row[11] or 0),
        "total": float(row[12] or 0)
        })
    except (ValueError, TypeError):
        continue  # Skip non-numeric rows

with open(os.path.join(DATA_DIR, 'boosts.json'), 'w') as f:
    json.dump(boosts, f, indent=1)
print(f"Boosts: {len(boosts)} entries written")

wb.close()
print("Done! Files written to tools/clash/data/")
